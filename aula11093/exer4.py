from openpyxl import Workbook

# Manipulando Excel
wb = Workbook()
ws = wb.active
ws.title = "Planilha Orçamentos"
ws['A1'] = "Produto."
ws['B1'] = "Valor"
ws['C1'] = "Quantidade"
ws['D1'] = "Total por Produto"
ws['E1'] = "Total Geral"
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10

linha = 2
nome = ""
while nome != 'FIM':
  nome = input("Informe o nome do Produto [FIM - Termina o programa]: ")
  if nome == 'FIM':
      continue

  valor = float(input("Qual o valor do produto?"))
  qtd = float(input("Qual a quantidade"))

  seq = str(linha-1)
  ws.cell(row=linha, column=1, value=nome)
  ws.cell(row=linha, column=2, value=valor)
  ws.cell(row=linha, column=3, value=qtd)

  tp = qtd * valor
  ws.cell(row=linha, column=4, value=tp)

  tg = tp
  ws.cell(row=linha, column=5, value=tg)

  linha += 1

wb.save("c:\\temp\\avaliacao.xlsx")
print ("Planilha salva em c:/temp/avaliacao.xlsx")
